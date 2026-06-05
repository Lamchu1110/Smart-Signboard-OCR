from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_PATH = BASE_DIR / "outputs" / "signboards.xlsx"

HEADERS = [
    "id",
    "image_name",
    "store_name",
    "category",
    "phone",
    "address",
    "services",
    "raw_text",
    "confidence",
    "created_at",
    "note",
]

COLUMN_WIDTHS = {
    "A": 10,
    "B": 28,
    "C": 32,
    "D": 20,
    "E": 18,
    "F": 48,
    "G": 38,
    "H": 60,
    "I": 14,
    "J": 20,
    "K": 32,
}


def style_worksheet(worksheet):
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    for column, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def ensure_workbook(path: Path = DEFAULT_OUTPUT_PATH):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook.active
        existing_headers = [cell.value for cell in worksheet[1]]
        if existing_headers != HEADERS:
            existing_rows = [
                {header: value for header, value in zip(existing_headers, row)}
                for row in worksheet.iter_rows(min_row=2, values_only=True)
            ]
            workbook.remove(worksheet)
            worksheet = workbook.create_sheet("signboards")
            worksheet.append(HEADERS)
            for index, row in enumerate(existing_rows, start=1):
                normalized = normalize_record(row)
                normalized["id"] = normalized["id"] or index
                worksheet.append([normalized[header] for header in HEADERS])
            style_worksheet(worksheet)
            workbook.save(path)
        return workbook, worksheet

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "signboards"
    worksheet.append(HEADERS)
    style_worksheet(worksheet)
    workbook.save(path)
    return workbook, worksheet


def next_record_id(worksheet):
    ids = []
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        try:
            ids.append(int(row[0]))
        except (TypeError, ValueError):
            continue
    return (max(ids) + 1) if ids else 1


def normalize_record(record: dict):
    return {
        "id": record.get("id", ""),
        "image_name": record.get("image_name", ""),
        "store_name": record.get("store_name") or record.get("business_name", ""),
        "category": record.get("category") or record.get("category_hint", ""),
        "phone": record.get("phone", ""),
        "address": record.get("address") or record.get("address_hint", ""),
        "services": record.get("services", ""),
        "raw_text": record.get("raw_text", ""),
        "confidence": record.get("confidence", ""),
        "created_at": record.get("created_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "note": record.get("note", ""),
    }


def append_signboard_record(record: dict, path: Path = DEFAULT_OUTPUT_PATH):
    workbook, worksheet = ensure_workbook(path)
    normalized = normalize_record(record)
    if not normalized["id"]:
        normalized["id"] = next_record_id(worksheet)
    worksheet.append([normalized[header] for header in HEADERS])
    style_worksheet(worksheet)
    row_number = worksheet.max_row
    workbook.save(path)
    workbook.close()
    return {
        "saved": True,
        "path": str(path),
        "row": row_number,
        "record": normalized,
    }


def read_signboard_records(path: Path = DEFAULT_OUTPUT_PATH):
    if not path.exists():
        return []

    workbook = load_workbook(path, read_only=True)
    try:
        worksheet = workbook.active
        rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append({header: value or "" for header, value in zip(HEADERS, row)})
        return rows
    finally:
        workbook.close()
